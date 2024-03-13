import openpyxl

file = "C:/Users/10708565/OneDrive - LTIMindtree/Documents/Data2.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Data1"]
rows=sheet.max_row
cols=sheet.max_column
print(cols)
print(rows)
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end=' ')
    print()